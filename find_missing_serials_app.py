'''import streamlit as st
import pandas as pd

def find_missing_serial_numbers(df, serial_column):
    # Ensure the serial column is sorted
    df = df.sort_values(by=[serial_column])

    # Create a range of expected serial numbers
    min_serial = df[serial_column].min()
    max_serial = df[serial_column].max()
    expected_serials = set(range(min_serial, max_serial + 1))

    # Find actual serial numbers present in the file
    actual_serials = set(df[serial_column])

    # Find missing serial numbers
    missing_serials = expected_serials - actual_serials

    if missing_serials:
        st.write(f"Missing serial numbers: {sorted(missing_serials)}")
    else:
        st.write("No missing serial numbers found.")

# Streamlit App
st.title("Find Missing Serial Numbers in Excel")

# File uploader widget
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

# When the file is uploaded
if uploaded_file:
    try:
        # Read the uploaded Excel file
        df = pd.read_excel(uploaded_file)
        
        # Input field to specify the serial number column name
        serial_column = st.text_input("Enter the name of the serial number column:")
        
        if serial_column:
            # Call the function to find missing serial numbers
            find_missing_serial_numbers(df, serial_column)
    except Exception as e:
        st.error(f"An error occurred: {e}")


'''

import streamlit as st
import pandas as pd
from openpyxl import load_workbook

def find_missing_serial_numbers(serial_numbers):
    # Ensure the serial numbers are sorted
    serial_numbers = sorted(serial_numbers)

    # Create a range of expected serial numbers
    min_serial = serial_numbers[0]
    max_serial = serial_numbers[-1]
    expected_serials = set(range(min_serial, max_serial + 1))

    # Find actual serial numbers present in the file
    actual_serials = set(serial_numbers)

    # Find missing serial numbers
    missing_serials = expected_serials - actual_serials

    if missing_serials:
        st.write(f"Missing serial numbers: {sorted(missing_serials)}")
    else:
        st.write("No missing serial numbers found.")

# Streamlit App
st.title("Find Missing Serial Numbers in Excel")

# File uploader widget
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

# When the file is uploaded
if uploaded_file:
    try:
        # Load the workbook
        workbook = load_workbook(uploaded_file, read_only=True)
        sheet = workbook.active
        
        # Input field to specify the serial number column name
        serial_column = st.text_input("Enter the name of the serial number column:")

        if serial_column:
            # Get the column names from the first row
            columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            if serial_column not in columns:
                st.error(f"Column '{serial_column}' not found in the file.")
            else:
                # Find the index of the serial number column
                serial_col_idx = columns.index(serial_column) + 1

                serial_numbers = []
                # Iterate through the rows and collect serial numbers
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    serial_number = row[serial_col_idx - 1]
                    if serial_number is not None:
                        serial_numbers.append(serial_number)

                # Call the function to find missing serial numbers
                if serial_numbers:
                    find_missing_serial_numbers(serial_numbers)
                else:
                    st.write("No serial numbers found in the specified column.")

    except Exception as e:
        st.error(f"An error occurred: {e}")
